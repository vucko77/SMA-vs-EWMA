#!/usr/bin/env python
# coding: utf-8

# In[ ]:


#Descriptivna - ama proveri za koi metriki e
import pandas as pd
from scipy.stats import shapiro, kstest, skew, kurtosis
from scipy.stats import norm

# Load your data
data = pd.read_excel('/Users/vladimirvuksanovikj/Downloads/Shani Anova/VeljkoStoev2.xlsx')  # Make sure to update this path

# Specify the metrics to analyze
metrics = [
    'Total Distance (m)', 'HMLD (m)', 'Total sprints distance (m)',
    'Sprints count', 'Top speed (km/h)', 'Accelerations', 'Decelerations'
]

# Prepare a DataFrame to store the descriptive statistics
desc_stats = pd.DataFrame(columns=['Metric', 'Min', 'Max', 'Mean', 'SD', 'Skew', 'Kurtosis', 'Shapiro-Wilk', 'Kolmogorov-Smirnov'])

# Calculate descriptive statistics and normality tests
for metric in metrics:
    series = data[metric].dropna()
    desc_stats = desc_stats.append({
        'Metric': metric,
        'Min': series.min(),
        'Max': series.max(),
        'Mean': series.mean(),
        'SD': series.std(),
        'Skew': skew(series),
        'Kurtosis': kurtosis(series),
        'Shapiro-Wilk': shapiro(series)[1],
        'Kolmogorov-Smirnov': kstest(series, 'norm', args=(series.mean(), series.std()))[1]
    }, ignore_index=True)

# Define a function to apply the color formatting
def highlight_non_normal(val):
    color = 'orange' if val < 0.05 else None
    return f'background-color: {color}'

# Apply the color formatting
styled_desc_stats = desc_stats.style.applymap(
    highlight_non_normal, subset=['Shapiro-Wilk', 'Kolmogorov-Smirnov']
)

# Print the styled DataFrame
print(styled_desc_stats)

# To save the styled DataFrame to an Excel file, you need xlsxwriter
excel_file = 'Descriptive_Statistics.xlsx'
styled_desc_stats.to_excel(excel_file, engine='xlsxwriter')

print(f'Descriptive statistics saved to {excel_file}.')


# In[4]:


pip install pandas scipy xlsxwriter


# In[2]:


pip install scikit-posthocs


# In[23]:


#ManWhitnet za eden igrach so site drugi

import pandas as pd
from scipy.stats import mannwhitneyu
# Man Witney U test megju eden player so MEAN od site drugi
# Reload the data
data_path = '/Users/vladimirvuksanovikj/Downloads/Shani Anova/VeljkoStoev2.xlsx'
data = pd.read_excel(data_path)

# Define the player of interest and the metrics to analyze
player_of_interest = 'Roberto Raichev'
metrics = [
    'Total Distance (m)', 'HMLD (m)', 'Total sprints distance (m)',
    'Sprints count', 'Top speed (km/h)', 'Accelerations', 'Decelerations'
]

# Prepare a DataFrame to store the test results
comparison_results = pd.DataFrame(columns=['Metric', 'U-Statistic', 'P-Value', 'Mean Veljko Jelenkovic', 'Mean Team'])

# Separate the data for Veljko Jelenkovic and the rest of the team
veljko_data = data[data['Player'] == player_of_interest]
team_data = data[data['Player'] != player_of_interest]

# Perform the Mann-Whitney U test for each metric
for metric in metrics:
    veljko_metric_data = veljko_data[metric].dropna()
    team_metric_data = team_data[metric].dropna()

    # Calculate the means for Veljko Jelenkovic and the rest of the team
    mean_veljko = veljko_metric_data.mean()
    mean_team = team_metric_data.mean()
    
    # Perform the Mann-Whitney U test if both groups have data
    if len(veljko_metric_data) > 0 and len(team_metric_data) > 0:
        u_stat, p_val = mannwhitneyu(veljko_metric_data, team_metric_data, alternative='two-sided')
        comparison_results = comparison_results.append({
            'Metric': metric,
            'U-Statistic': u_stat,
            'P-Value': p_val,
            'Mean Veljko Jelenkovic': mean_veljko,
            'Mean Team': mean_team
        }, ignore_index=True)
    else:
        comparison_results = comparison_results.append({
            'Metric': metric,
            'U-Statistic': 'N/A',
            'P-Value': 'N/A',
            'Mean Veljko Jelenkovic': mean_veljko if len(veljko_metric_data) > 0 else 'N/A',
            'Mean Team': mean_team if len(team_metric_data) > 0 else 'N/A'
        }, ignore_index=True)

# Display the comparison results
comparison_results


# In[30]:


#ManWhitnet za eden igrach so site drugi, ne se vkluceni drugite povredeni

import pandas as pd
from scipy.stats import mannwhitneyu
# Man Witney U test megju eden player so MEAN od drugite ama ne se vkluceni povredenite
# Load the data
data_path = '/Users/vladimirvuksanovikj/Downloads/Shani Anova/VeljkoStoev2.xlsx'  # Ensure the path is correct
data = pd.read_excel(data_path)

# Define the player of interest and the metrics to analyze
player_of_interest = 'Toni Tasev'
exclude_players = ['Veljko Jelenkovic', 'Emil Stoev', 'Galin Ivanov', 'Denislav Aleksandrov']
metrics = [
    'Total Distance (m)', 'HMLD (m)', 'Total sprints distance (m)',
    'Sprints count', 'Top speed (km/h)', 'Accelerations', 'Decelerations'
]

# Prepare a DataFrame to store the test results
comparison_results = pd.DataFrame(columns=['Metric', 'U-Statistic', 'P-Value', 'Mean Veljko Jelenkovic', 'Mean Others'])

# Separate the data for Veljko Jelenkovic and the others excluding specific players
veljko_data = data[data['Player'] == player_of_interest]
others_data = data[(data['Player'] != player_of_interest) & (~data['Player'].isin(exclude_players))]

# Perform the Mann-Whitney U test for each metric
for metric in metrics:
    veljko_metric_data = veljko_data[metric].dropna()
    others_metric_data = others_data[metric].dropna()

    # Calculate the means for Veljko Jelenkovic and the modified others group
    mean_veljko = veljko_metric_data.mean()
    mean_others = others_metric_data.mean()
    
    # Perform the Mann-Whitney U test if both groups have data
    if len(veljko_metric_data) > 0 and len(others_metric_data) > 0:
        u_stat, p_val = mannwhitneyu(veljko_metric_data, others_metric_data, alternative='two-sided')
        comparison_results = comparison_results.append({
            'Metric': metric,
            'U-Statistic': u_stat,
            'P-Value': p_val,
            'Mean Veljko Jelenkovic': mean_veljko,
            'Mean Others': mean_others
        }, ignore_index=True)
    else:
        comparison_results = comparison_results.append({
            'Metric': metric,
            'U-Statistic': 'N/A',
            'P-Value': 'N/A',
            'Mean Veljko Jelenkovic': mean_veljko if len(veljko_metric_data) > 0 else 'N/A',
            'Mean Others': mean_others if len(others_metric_data) > 0 else 'N/A'
        }, ignore_index=True)
# Display the comparison results
comparison_results
# Export the comparison results to an Excel file
#output_excel_path = '/mnt/data/Comparison_Results_Veljko_Exclude.xlsx'
#comparison_results.to_excel(output_excel_path, index=False)
#print(f"The comparison results have been saved to {output_excel_path}.")


# In[19]:


#Kruskal Wallis H test da proveram dali celiot sistem na igraci vo site denovi kakvo p-level ima 
#/Users/vladimirvuksanovikj/Downloads/Shani Anova/VeljkoStoev2.xlsx

import pandas as pd
from scipy.stats import kruskal
from IPython.display import display

# Load the data
data_path = '/Users/vladimirvuksanovikj/Downloads/Shani Anova/VeljkoStoev2.xlsx'  # Adjust the path as needed
data = pd.read_excel(data_path)

# Define the metrics to analyze
metrics = [
    'Total Distance (m)', 
    'Distance 3+4+5 (m)',  # Assuming the column name is exactly this, without parentheses
    'HMLD (m)', 
    'Accelerations', 
    'Decelerations'
]

# Prepare a DataFrame to store the Kruskal-Wallis test results
kw_results = pd.DataFrame(columns=['Metric', 'H-Statistic', 'P-Value', 'Is Significant'])

# Perform the Kruskal-Wallis H test for each metric
for metric in metrics:
    # Check if the metric exists in the DataFrame to prevent KeyError
    if metric in data.columns:
        # Prepare the list of data arrays for the test, one array per player
        grouped_data = [group.dropna().values for name, group in data.groupby('Player')[metric] if len(group.dropna()) > 0]

        # Perform the Kruskal-Wallis H test if there is enough data
        if len(grouped_data) > 1:
            h_statistic, p_value = kruskal(*grouped_data)
            kw_results = kw_results.append({
                'Metric': metric,
                'H-Statistic': h_statistic,
                'P-Value': p_value,
                'Is Significant': p_value < 0.05  # Assuming a significance level of 0.05
            }, ignore_index=True)
        else:
            kw_results = kw_results.append({
                'Metric': metric,
                'H-Statistic': 'N/A',
                'P-Value': 'N/A',
                'Is Significant': 'N/A'
            }, ignore_index=True)
    else:
        print(f"Metric {metric} not found in the dataset.")

# Display the results table directly in the Jupyter Notebook
display(kw_results)


# In[4]:


pip install scikit-posthocs


# In[21]:


# Dunn posthoc test , doagja posle Kruskal Wallish H Test, ama i vo ovoj kod pravi Kruskal Walis prvin
#'/Users/vladimirvuksanovikj/Downloads/Shani Anova/VeljkoStoev2.xlsx'

import pandas as pd
from scipy.stats import kruskal
import scikit_posthocs as sp
from IPython.display import display

# Load the data
data_path = '/Users/vladimirvuksanovikj/Downloads/Shani Anova/VeljkoStoev2.xlsx'
data = pd.read_excel(data_path)

# Define the metrics to analyze
metrics = [
    'Total Distance (m)',
    'Distance 3+4+5 (m)',
    'HMLD (m)',
    'Accelerations',
    'Decelerations'
]

# Process each metric
for metric in metrics:
    print(f"Processing metric: {metric}")
    groups = data.groupby('Player')[metric].apply(list).to_dict()
    grouped_data = [g for g in groups.values() if len(g) > 0]

    if len(grouped_data) > 1:
        k_stat, p_val = kruskal(*grouped_data)
        if p_val < 0.05:
            flat_data = [item for sublist in grouped_data for item in sublist]
            group_labels = [k for k, v in groups.items() for _ in v]

            posthoc_data = pd.DataFrame({'Value': flat_data, 'Group': group_labels})
            p_matrix = sp.posthoc_dunn(posthoc_data, val_col='Value', group_col='Group', p_adjust='bonferroni')

            # Apply styling for significant p-values for notebook display
            styled_p_matrix = p_matrix.style.applymap(lambda x: 'background-color: #FFA07A' if x < 0.05 else '')
            display(styled_p_matrix)

            # Save to Excel with conditional formatting
            output_excel_path = f'/Users/vladimirvuksanovikj/Downloads/Shani Anova/Dunn_Post_Hoc_{metric.replace(" ", "_")}.xlsx'
            writer = pd.ExcelWriter(output_excel_path, engine='xlsxwriter')
            p_matrix.to_excel(writer, sheet_name='Post_Hoc_Results')
            workbook = writer.book
            worksheet = writer.sheets['Post_Hoc_Results']
            format_red = workbook.add_format({'bg_color': '#FFA07A'})

            # Excel conditional formatting
            worksheet.conditional_format('B2:Z1000', {  # Adjust the range according to your actual data range
                'type': 'cell',
                'criteria': '<',
                'value': 0.05,
                'format': format_red
            })
            writer.save()
            print(f"Results for {metric} saved to {output_excel_path}")
        else:
            print(f"No significant differences found for {metric} via Kruskal-Wallis (p={p_val}).")
    else:
        print(f"Not enough data to perform Kruskal-Wallis test on {metric}.")


# In[18]:


# ManWhitney matrix na igraci za SITE metriki
#/Users/vladimirvuksanovikj/Downloads/Shani Anova/VeljkoStoev2.xlsx
#/Users/vladimirvuksanovikj/Downloads/Shani Anova/Dunn_Post_Hoc_Results.xlsx
#Total Distance (m)
#Distance 3+4+5 (m)
#HMLD (m)
#Accelerations
#Decelerations

import pandas as pd
import numpy as np
from scipy.stats import mannwhitneyu
import xlsxwriter
from itertools import combinations

# Load the data
data_path = '/Users/vladimirvuksanovikj/Downloads/Shani Anova/VeljkoStoev2.xlsx'
data = pd.read_excel(data_path)

# Define the metrics to analyze
metrics = [
    'Total Distance (m)', 
    'Distance 3+4+5 (m)', 
    'HMLD (m)', 
    'Accelerations', 
    'Decelerations'
]

# Iterate through each metric and perform the analysis
for metric in metrics:
    # Get unique group names
    groups = data['Player'].unique()

    # Initialize a DataFrame to store the p-values in a matrix form
    p_value_matrix = pd.DataFrame(index=groups, columns=groups, dtype=float)

    # Conduct Mann-Whitney U tests between all pairs of groups and store the results in the matrix
    for group1, group2 in combinations(groups, 2):
        data1 = data[data['Player'] == group1][metric].dropna()
        data2 = data[data['Player'] == group2][metric].dropna()

        if not data1.empty and not data2.empty:
            _, p = mannwhitneyu(data1, data2, alternative='two-sided')
            p_value_matrix.at[group1, group2] = p
            p_value_matrix.at[group2, group1] = p  # Symmetric matrix

    # Save the p-value matrix to an Excel file with conditional formatting
    output_excel_path = f'/Users/vladimirvuksanovikj/Downloads/Shani Anova/P_Value_Matrix_{metric.replace(" ", "_")}.xlsx'
    writer = pd.ExcelWriter(output_excel_path, engine='xlsxwriter')
    p_value_matrix.to_excel(writer, sheet_name='P_Value_Matrix')

    # Get the workbook and the worksheet for formatting
    workbook = writer.book
    worksheet = writer.sheets['P_Value_Matrix']

    # Define a format for significant p-values: red text on a light red background
    format_significant = workbook.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})

    # Apply conditional formatting to highlight significant p-values
    row_count, col_count = p_value_matrix.shape
    worksheet.conditional_format(1, 1, row_count, col_count, {
        'type': 'cell',
        'criteria': '<',
        'value': 0.05,  # Assuming a significance level of 0.05
        'format': format_significant
    })

    writer.save()
    print(f"P-value matrix for {metric} with highlighted significance saved to {output_excel_path}.")


# In[17]:


#Proveruva ime na koloni dali se tocni vo excelo i vo kodot
import pandas as pd

# Load the data
data_path = '/Users/vladimirvuksanovikj/Downloads/Shani Anova/VeljkoStoev2.xlsx'
data = pd.read_excel(data_path)

# Print all column names to check for the correct names
print(data.columns.tolist())


# In[23]:


#Broi, count kolku metrici se znacajni za sekoj igrac, vo mannwhitneyu post hoc test
import pandas as pd
from scipy.stats import mannwhitneyu
import xlsxwriter
from itertools import combinations

# Load the data
data_path = '/Users/vladimirvuksanovikj/Downloads/Shani Anova/VeljkoStoev2.xlsx'  # Adjust the path to your Excel file
data = pd.read_excel(data_path)

# Define the metrics to analyze
metrics = [
    'Total Distance (m)', 
    'Distance 3+4+5 (m)', 
    'HMLD (m)', 
    'Accelerations', 
    'Decelerations'
]

# Initialize a dictionary to collect significant counts for each player
significant_counts = {metric: pd.Series(dtype=int) for metric in metrics}

# Process each metric
for metric in metrics:
    print(f"Processing {metric}")
    # Get unique groups (players)
    groups = data['Player'].unique()
    
    # Collect all p-values in a matrix form (optional, for visualization or other analysis)
    p_values = pd.DataFrame(index=groups, columns=groups, dtype=float)
    
    # Perform Mann-Whitney U tests between all pairs of groups
    for group1, group2 in combinations(groups, 2):
        data1 = data[data['Player'] == group1][metric].dropna()
        data2 = data[data['Player'] == group2][metric].dropna()

        if len(data1) > 0 and len(data2) > 0:
            _, p_value = mannwhitneyu(data1, data2, alternative='two-sided')
            p_values.at[group1, group2] = p_value
            p_values.at[group2, group1] = p_value

            # Check if the result is significant
            if p_value < 0.05:
                significant_counts[metric][group1] = significant_counts[metric].get(group1, 0) + 1
                significant_counts[metric][group2] = significant_counts[metric].get(group2, 0) + 1

# Convert dictionary to DataFrame for better manipulation and output
summary_df = pd.DataFrame(significant_counts).fillna(0)  # Replace NaN with 0 for non-significant counts

# Output the results to an Excel file
output_excel_path = '/Users/vladimirvuksanovikj/Downloads/Shani Anova/Significant_MW_Counts.xlsx'
with pd.ExcelWriter(output_excel_path, engine='xlsxwriter') as writer:
    summary_df.to_excel(writer, index_label='Player', sheet_name='Significant_Counts')

print(f"Results have been saved to {output_excel_path}.")


# In[24]:


#Broi, count kolku metrici se znacajni za sekoj igrac, vo DUNN post hoc test

#What the Script Does:
#Kruskal-Wallis Test: It is performed for each metric to see if there are any overall differences among groups. If significant differences are found (p < 0.05), Dunn's post hoc test is performed.
#Dunn's Post Hoc Test: This test is conducted to compare every pair of groups for each metric that showed significant results in the Kruskal-Wallis test. Dunn’s test corrects for multiple comparisons using the Bonferroni method.
#Counting Significant Comparisons: For each player, the script counts how many times they were part of a pairwise comparison that resulted in a significant p-value (p < 0.05). It subtracts 1 from each total count to exclude self-comparisons (a player compared with themselves, which isn’t actually computed but hypothetically would not be significant).
#Result Output: The counts of significant comparisons for each player are then saved to an Excel file. This gives you a summary of how each player's measurements significantly differ from others' across the metrics tested.


import pandas as pd
from scipy.stats import kruskal
import scikit_posthocs as sp

# Load the data
data_path = '/Users/vladimirvuksanovikj/Downloads/Shani Anova/VeljkoStoev2.xlsx'  # Adjust the path to your Excel file
data = pd.read_excel(data_path)

# Define the metrics to analyze
metrics = [
    'Total Distance (m)', 
    'Distance 3+4+5 (m)',
    'HMLD (m)', 
    'Accelerations', 
    'Decelerations'
]

# Initialize a dictionary to store the results DataFrame for each metric
results_summary = pd.DataFrame()

for metric in metrics:
    print(f"Processing {metric}")
    # Group data by 'Player' and gather their metric data, ensuring no empty groups
    groups = data.groupby('Player')[metric].apply(list).to_dict()
    grouped_data = [groups[player] for player in groups if len(groups[player]) > 0]

    if len(grouped_data) > 1:
        stat, p_val = kruskal(*grouped_data)
        if p_val < 0.05:
            # Perform Dunn's post hoc test
            flat_data = [item for sublist in grouped_data for item in sublist]
            group_labels = [player for player in groups for _ in groups[player]]

            posthoc_data = pd.DataFrame({'Value': flat_data, 'Group': group_labels})
            posthoc_result = sp.posthoc_dunn(posthoc_data, val_col='Value', group_col='Group', p_adjust='bonferroni')

            # Count significant p-values for each player
            sig_counts = (posthoc_result < 0.05).sum()
            sig_counts.name = metric  # Rename the Series for easier concatenation
            results_summary = pd.concat([results_summary, sig_counts], axis=1, sort=False)

# Replace negative values with 0
results_summary[results_summary < 0] = 0

# Output the results to an Excel file
output_excel_path = '/Users/vladimirvuksanovikj/Downloads/Shani Anova/Significant_P_Values_Count.xlsx'
results_summary.to_excel(output_excel_path, index_label='Player')

print(f"Results have been saved to {output_excel_path}.")


# In[27]:


#compare the counts of significant p-values from Dunn's post hoc test and the Mann-Whitney U test 
#and calculate the percentage difference

import pandas as pd
from scipy.stats import kruskal, mannwhitneyu
import scikit_posthocs as sp
from itertools import combinations
from scipy.stats import ttest_rel, wilcoxon

# Load the data
data_path = '/Users/vladimirvuksanovikj/Downloads/Shani Anova/VeljkoStoev2.xlsx'  # Adjust the path to your Excel file
data = pd.read_excel(data_path)

# Define the metrics to analyze
metrics = [
    'Total Distance (m)', 
    'Distance 3+4+5 (m)',
    'HMLD (m)', 
    'Accelerations', 
    'Decelerations'
]

# Initialize lists to store results
dunn_counts = []
mw_counts = []

for metric in metrics:
    print(f"Processing {metric}")
    # Group data by 'Player' and gather their metric data, ensuring no empty groups
    groups = data.groupby('Player')[metric].apply(list).to_dict()
    grouped_data = [groups[player] for player in groups if len(groups[player]) > 0]

    if len(grouped_data) > 1:
        stat, p_val = kruskal(*grouped_data)
        if p_val < 0.05:
            # Perform Dunn's post hoc test
            flat_data = [item for sublist in grouped_data for item in sublist]
            group_labels = [player for player in groups for _ in groups[player]]

            posthoc_data = pd.DataFrame({'Value': flat_data, 'Group': group_labels})
            posthoc_result = sp.posthoc_dunn(posthoc_data, val_col='Value', group_col='Group', p_adjust='bonferroni')

            # Count significant p-values for Dunn's test
            sig_counts_dunn = (posthoc_result < 0.05).sum() - 1  # Subtract 1 to exclude self-comparison
            dunn_counts.append(sig_counts_dunn.sum())  # Sum the counts across all comparisons

            # Perform Mann-Whitney U test
            mw_result = pd.DataFrame(columns=['Player', 'P-Value'])
            for group1, group2 in combinations(groups, 2):
                data1 = data[data['Player'] == group1][metric].dropna()
                data2 = data[data['Player'] == group2][metric].dropna()
                mw_stat, mw_p_val = mannwhitneyu(data1, data2, alternative='two-sided')
                mw_result = mw_result.append({'Player': f"{group1} vs {group2}", 'P-Value': mw_p_val}, ignore_index=True)

            # Count significant p-values for Mann-Whitney U test
            sig_counts_mw = (mw_result['P-Value'] < 0.05).sum()
            mw_counts.append(sig_counts_mw)

# Calculate percentage difference
percentage_difference = [(d - m) / m * 100 for d, m in zip(dunn_counts, mw_counts)]


# Perform paired t-test
t_stat, t_p_val = ttest_rel(dunn_counts, mw_counts)

# Perform Wilcoxon signed-rank test
w_stat, w_p_val = wilcoxon(dunn_counts, mw_counts)

# printi tuka vo Jupeter dali ima stat znacajna razlika megju dvata counta
print(f"Paired t-test p-value: {t_p_val}")
print(f"Wilcoxon signed-rank test p-value: {w_p_val}")


# Create DataFrame with results
results_df = pd.DataFrame({'Dunn Count': dunn_counts, 'Mann-Whitney Count': mw_counts, '% Difference': percentage_difference}, index=metrics)

# Output the results to an Excel file
output_excel_path = '/Users/vladimirvuksanovikj/Downloads/Shani Anova/Significant_P_Values_Comparison.xlsx'
results_df.to_excel(output_excel_path, index_label='Metric')

print(f"Results have been saved to {output_excel_path}.")


# In[1]:


#This script will count the significant p-values of Dunn's test for each player across the specified metrics, 
#calculate the total number of Dunn tests done for each player, and compute the percentage difference between 
#significant Dunn tests and total Dunn tests for each player. Finally, it saves the results in an Excel table.

import pandas as pd
import numpy as np
import scikit_posthocs as sp
from collections import Counter

# Load the data
data_path = '/Users/vladimirvuksanovikj/Downloads/Shani Anova/VeljkoStoev2.xlsx'  # Adjust the path to your Excel file
data = pd.read_excel(data_path)

# Define the metrics to analyze
metrics = ['Total Distance (m)', 'Distance 3+4+5 (m)', 'HMLD (m)', 'Accelerations', 'Decelerations']

# Initialize dictionaries to store results
sig_counts = Counter()
total_counts = Counter()

# Loop through each metric
for metric in metrics:
    # Group data by 'Player' and gather their metric data
    groups = data.groupby('Player')[metric].apply(list).to_dict()
    
    # Flatten the grouped data and remove NaN values
    flat_data = [item for sublist in groups.values() for item in sublist if not pd.isnull(item)]
    
    # Perform Dunn's test
    dunn_result = sp.posthoc_dunn(pd.DataFrame({'Value': flat_data, 'Player': [player for player in groups for _ in groups[player]]}),
                                  val_col='Value', group_col='Player', p_adjust='bonferroni')
    
    # Count significant p-values for each player
    for player in groups.keys():
        p_vals = dunn_result.loc[player].drop(player).values
        sig_counts[player] += np.sum(p_vals < 0.05)
        total_counts[player] += len(p_vals)

# Convert counts to DataFrame
results_df = pd.DataFrame({'Player': list(sig_counts.keys()),
                           'Significant Counts': list(sig_counts.values()),
                           'Total Counts': list(total_counts.values())})

# Calculate percentage difference
results_df['Percentage Difference'] = (results_df['Significant Counts'] / results_df['Total Counts']) * 100

# Save results to Excel
output_excel_path = '/Users/vladimirvuksanovikj/Downloads/Shani Anova/results.xlsx'  # Adjust the path for output Excel file
results_df.to_excel(output_excel_path, index=False)

print(f"Results have been saved to {output_excel_path}.")


# In[2]:


#This script will count the significant p-values of the Mann-Whitney U test for each player across the specified 
#metrics, calculate the total number of Mann-Whitney U tests done for each player, and compute the percentage 
#difference between significant Mann-Whitney U tests and total Mann-Whitney U tests for each player. 
#Finally, it saves the results in an Excel table.


import pandas as pd
import numpy as np
from itertools import combinations
from scipy.stats import mannwhitneyu

# Load the data
data_path = '/Users/vladimirvuksanovikj/Downloads/Shani Anova/VeljkoStoev2.xlsx'  # Adjust the path to your Excel file
data = pd.read_excel(data_path)

# Define the metrics to analyze
metrics = ['Total Distance (m)', 'Distance 3+4+5 (m)', 'HMLD (m)', 'Accelerations', 'Decelerations']

# Initialize dictionaries to store results
sig_counts = {}
total_counts = {}

# Loop through each metric
for metric in metrics:
    # Group data by 'Player' and gather their metric data
    groups = data.groupby('Player')[metric].apply(list).to_dict()
    
    # Perform Mann-Whitney U test between all pairs of groups
    for group1, group2 in combinations(groups.keys(), 2):
        data1 = data[data['Player'] == group1][metric].dropna()
        data2 = data[data['Player'] == group2][metric].dropna()
        
        stat, p_val = mannwhitneyu(data1, data2)
        
        # Update count of significant p-values for group1
        if group1 in sig_counts:
            sig_counts[group1] += 1 if p_val < 0.05 else 0
        else:
            sig_counts[group1] = 1 if p_val < 0.05 else 0
        
        # Update count of significant p-values for group2
        if group2 in sig_counts:
            sig_counts[group2] += 1 if p_val < 0.05 else 0
        else:
            sig_counts[group2] = 1 if p_val < 0.05 else 0
            
        # Update total counts for both groups
        total_counts[group1] = total_counts.get(group1, 0) + 1
        total_counts[group2] = total_counts.get(group2, 0) + 1

# Convert counts to DataFrame
results_df = pd.DataFrame({'Player': list(sig_counts.keys()),
                           'Significant Counts': list(sig_counts.values()),
                           'Total Counts': list(total_counts.values())})

# Calculate percentage difference
results_df['Percentage Difference'] = (results_df['Significant Counts'] / results_df['Total Counts']) * 100

# Save results to Excel
output_excel_path = '/Users/vladimirvuksanovikj/Downloads/Shani Anova/results2.xlsx'  # Adjust the path for output Excel file
results_df.to_excel(output_excel_path, index=False)

print(f"Results have been saved to {output_excel_path}.")


# In[14]:


#This code calculates ACWR for SMA and ACWR for EWMA, then calculates the correlation between them for each player, 
#for each date. It also highlights rows where any ACWR value is above 1.5 and exports the results to an Excel file. 
#/Users/vladimirvuksanovikj/Downloads/Shani Anova/SMA_vs_EWMA.xlsx

import pandas as pd

import pandas as pd
import numpy as np

# Load the data
data_path = '/Users/vladimirvuksanovikj/Downloads/Shani Anova/SMA_vs_EWMA.xlsx'  # Adjust the path to your Excel file
data = pd.read_excel(data_path)

# Filter the data for Denislav Aleksandrov
denislav_data = data[data['Player'] == 'Denislav Aleksandrov'].copy()  # Make a copy to avoid the warning

# Calculate ACWR for Total Distance (m) using Exponentially Weighted Moving Average (EWMA)
denislav_data.loc[:, 'ACWR_EWMA'] = denislav_data['Total Distance (m)'].ewm(span=7, min_periods=1).mean()                                     / denislav_data['Total Distance (m)'].ewm(span=28, min_periods=1).mean()

# Drop rows with NaN values (if any)
denislav_data = denislav_data.dropna()

# Select the desired columns
output_data = denislav_data[['Player', 'Day', 'Total Distance (m)', 'ACWR_EWMA']]

# Export the result to Excel
output_excel_path = '/Users/vladimirvuksanovikj/Downloads/Shani Anova/output.xlsx'  # Adjust the path for the output Excel file
output_data.to_excel(output_excel_path, index=False)
print(f"Results have been saved to {output_excel_path}.")


# In[13]:


print(data.columns)


# In[21]:


#All players, All metrics-  EWMA
import pandas as pd
import numpy as np

# Load the data
data_path = '/Users/vladimirvuksanovikj/Downloads/Shani Anova/SMA_vs_EWMA.xlsx'  # Adjust the path to your Excel file
data = pd.read_excel(data_path)

# Define the metrics to analyze
metrics = [
    'Total Distance (m)', 
    'Distance(3+4+5) (m)',
    'HMLD (m)', 
    'Accelerations', 
    'Decelerations'
]

# Initialize an empty DataFrame to store the results
results = pd.DataFrame()

# Iterate over each player
for player in data['Player'].unique():
    # Filter the data for the current player
    player_data = data[data['Player'] == player].copy()

    # Calculate ACWR for each metric based on EWMA
    for metric in metrics:
        acute_ewma = player_data[metric].ewm(span=7, min_periods=1).mean()
        chronic_ewma = player_data[metric].ewm(span=28, min_periods=1).mean()
        acwr_ewma_col = f'ACWR_{metric}_EWMA'
        player_data[acwr_ewma_col] = acute_ewma / chronic_ewma

    # Drop rows with NaN values (if any)
    player_data = player_data.dropna()

    # Select the desired columns
    player_results = player_data[['Player', 'Day'] + [f'ACWR_{metric}_EWMA' for metric in metrics]]

    # Append the results to the overall DataFrame
    results = pd.concat([results, player_results])

# Export the result to Excel
output_excel_path = '/Users/vladimirvuksanovikj/Downloads/Shani Anova/EWMA_All_All.xlsx'  # Adjust the path for the output Excel file
results.to_excel(output_excel_path, index=False)
print(f"Results have been saved to {output_excel_path}.")


# In[22]:


#based on RScript kodot
import pandas as pd
import numpy as np
from scipy.signal import savgol_filter

# Load the data from the Excel file
data = pd.read_excel('/Users/vladimirvuksanovikj/Downloads/Shani Anova/SMA_vs_EWMA.xlsx')

# Function to calculate exponential weighted moving average (EWMA) with non-zero check
def calc_ewma_nonzero(series, span):
    nonzero_indices = series[series != 0].index
    ewma_values = pd.Series(np.nan, index=series.index)
    ewma_values[nonzero_indices] = series[nonzero_indices].ewm(span=span, min_periods=1).mean()
    return ewma_values

# Perform calculations
data['rolling_avg_pracma_e7'] = calc_ewma_nonzero(data['Distance(3+4+5) (m)'], 7)
data['rolling_avg_pracma_e28'] = calc_ewma_nonzero(data['Distance(3+4+5) (m)'], 28)
data['rolling_avg_hmld_e7'] = calc_ewma_nonzero(data['HMLD (m)'], 7)
data['rolling_avg_hmld_e28'] = calc_ewma_nonzero(data['HMLD (m)'], 28)
data['rolling_avg_total_distance_e7'] = calc_ewma_nonzero(data['Total Distance (m)'], 7)
data['rolling_avg_total_distance_e28'] = calc_ewma_nonzero(data['Total Distance (m)'], 28)
data['rolling_avg_accelerations_e7'] = calc_ewma_nonzero(data['Accelerations'], 7)
data['rolling_avg_accelerations_e28'] = calc_ewma_nonzero(data['Accelerations'], 28)
data['rolling_avg_decelerations_e7'] = calc_ewma_nonzero(data['Decelerations'], 7)
data['rolling_avg_decelerations_e28'] = calc_ewma_nonzero(data['Decelerations'], 28)

# Calculate ACWR
data['pracma_e_acwr'] = data['rolling_avg_pracma_e7'] / data['rolling_avg_pracma_e28'].replace(0, np.nan)
data['hmld_e_acwr'] = data['rolling_avg_hmld_e7'] / data['rolling_avg_hmld_e28'].replace(0, np.nan)
data['total_distance_e_acwr'] = data['rolling_avg_total_distance_e7'] / data['rolling_avg_total_distance_e28'].replace(0, np.nan)
data['accelerations_e_acwr'] = data['rolling_avg_accelerations_e7'] / data['rolling_avg_accelerations_e28'].replace(0, np.nan)
data['decelerations_e_acwr'] = data['rolling_avg_decelerations_e7'] / data['rolling_avg_decelerations_e28'].replace(0, np.nan)

# Select only necessary columns
data_pracma = data[['Player', 'Day', 'Distance(3+4+5) (m)', 'HMLD (m)', 'Total Distance (m)', 'Accelerations', 'Decelerations', 'pracma_e_acwr', 'hmld_e_acwr', 'total_distance_e_acwr', 'accelerations_e_acwr', 'decelerations_e_acwr']]

# Export the result to an Excel file
data_pracma.to_excel('/Users/vladimirvuksanovikj/Downloads/Shani Anova/Pyhton_Rscript.xlsx', index=False)


# In[24]:


pip install pandas openpyxl


# In[26]:


pip install pandas openpyxl scipy


# In[30]:


#Pearson correlation—such as the p-value, 
#Calculates Pearson Correlation Coefficient and P-value for each pair of columns specified.
#Calculates (𝑟2)to determine how much variance in one variable is explained by the other.
#Calculates the 95% Confidence Interval using the Fisher Z transformation, providing a range within which the true correlation coefficient is likely to fall with 95% certainty.
#Stores all results in a pandas DataFrame and outputs to an Excel file.
#Sample size as new column. 

import pandas as pd
from scipy.stats import pearsonr
import numpy as np

# Path to the source Excel file
source_file = "/Users/vladimirvuksanovikj/Downloads/Shani Anova/EXCEL Za trudot/Power bi/korelacija od power Bi.xlsx"
# Path to the destination Excel file
destination_file = "/Users/vladimirvuksanovikj/Downloads/Shani Anova/EXCEL Za trudot/Power bi/korelacija.xlsx"

# Read the Excel file
df = pd.read_excel(source_file)

# Define the column pairs for correlation calculation
pairs = [
    ("EWMA_ACWR DZ_3+4+5", "DZ 3+4+5_ACWR"),
    ("EWMA_ACWR HMLD", "HMLD_ACWR"),
    ("EWMA_ACWR TD", "Total D_ACWR"),
    ("EWMA_ACWR ACC", "ACC_ACWR"),
    ("EWMA_ACWR DEC", "DEC_ACWR")
]

# Calculate correlations, p-values, r-squared, confidence intervals, and store them in a list
results = []
for col1, col2 in pairs:
    if df[col1].notna().all() and df[col2].notna().all():
        corr, p_value = pearsonr(df[col1], df[col2])
        r_squared = corr**2
        sample_size = len(df[col1].dropna())
        se = 1 / np.sqrt(sample_size - 3)
        ci_lower = np.tanh(np.arctanh(corr) - 1.96 * se)
        ci_upper = np.tanh(np.arctanh(corr) + 1.96 * se)
        results.append({
            'Column Pair': f'{col1} & {col2}',
            'Correlation': corr,
            'P-value': p_value,
            'Coefficient of Determination (r^2)': r_squared,
            '95% Confidence Interval Lower': ci_lower,
            '95% Confidence Interval Upper': ci_upper,
            'Sample Size': sample_size
        })

# Create a DataFrame from the results
result_df = pd.DataFrame(results)

# Write the result DataFrame to a new Excel file
result_df.to_excel(destination_file, index=False)

print(f"Correlation results have been saved to '{destination_file}'.")


# In[34]:


pip install matplotlib pandas openpyxl numpy


# In[12]:


#Script for Creating Scatter Plots не ги преклопува по старо

import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import os
from io import BytesIO
import numpy as np

# Path to the source Excel file
source_file = "/Users/vladimirvuksanovikj/Downloads/Shani Anova/Abstrakt 2/Excel/Glavna za vo Python.xlsx"
# Path to the destination Excel file
destination_file = "/Users/vladimirvuksanovikj/Downloads/Shani Anova/Abstrakt 2/Excel/ScatterPlots.xlsx"

# Read the Excel file
df = pd.read_excel(source_file)

# Define the column pairs for correlation calculation
pairs = [
    ("DZ345_EWMA", "DZ345_SMA"),
    ("HMLD_EWMA", "HMLD_SMA"),
    ("TD_EWMA", "TD_SMA"),
    ("ACC_EWMA", "ACC_SMA"),
    ("DEC_EWMA", "DEC_SMA")
]

# Create a new Excel workbook and select the active sheet
wb = Workbook()
ws = wb.active

# Create scatter plots for each pair of columns
for index, (col1, col2) in enumerate(pairs, 1):
    plt.figure(figsize=(8, 6))

    # Color the dots based on column groups
    plt.scatter(df[col1], df[col2], alpha=0.5, c=np.where(df[col1] >= df[col2], 'red', 'blue'), label=f'{col1} (red) vs. {col2} (blue)')
    plt.title(f'Scatter Plot between {col1} and {col2}')
    plt.xlabel(col1)
    plt.ylabel(col2)
    plt.legend()
    plt.grid(True)

    # Save the plot to a bytes object and then to an image file
    img_data = BytesIO()
    plt.savefig(img_data, format='png')
    img_data.seek(0)
    img = Image(img_data)
    # Add image to the Excel sheet
    cell_location = f'A{1 + (index - 1) * 20}'  # Adjust cell location for each image
    ws.add_image(img, cell_location)
    plt.close()

# Save the workbook with the images
wb.save(destination_file)
print(f"All scatter plots have been saved to '{destination_file}'.")


# In[23]:


#Scater ovoj e finalen gi mrdna (offset/ da ne se preklopuvaat)
import pandas as pd
import matplotlib.pyplot as plt
from scipy.stats import spearmanr
from scipy.optimize import curve_fit
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from io import BytesIO
import numpy as np

# Function to fit a polynomial curve
def quadratic_fit(x, a, b, c):
    return a * x**2 + b * x + c

# Source and destination file paths
source_file = "/Users/vladimirvuksanovikj/Downloads/Shani Anova/Abstrakt 2/Excel/Glavna za vo Python.xlsx"
destination_file = "/Users/vladimirvuksanovikj/Downloads/Shani Anova/Abstrakt 2/Excel/ScatterPlots.xlsx"

# Read the Excel file
df = pd.read_excel(source_file)

# Define the column pairs for correlation calculation
pairs = [
    ("DZ345_EWMA", "DZ345_SMA"),
    ("HMLD_EWMA", "HMLD_SMA"),
    ("TD_EWMA", "TD_SMA"),
    ("ACC_EWMA", "ACC_SMA"),
    ("DEC_EWMA", "DEC_SMA")
]

# Create a new Excel workbook and select the active sheet
wb = Workbook()
ws = wb.active

# Create scatter plots for each pair of columns
for index, (col1, col2) in enumerate(pairs, 1):
    plt.figure(figsize=(8, 6))
    
    # Adjusting the y-values of col2 for better visibility
    adjusted_y = df[col2] + 0.5  # Shift y-values of col2 up by 0.5

    # Scatter plot settings for each variable
    plt.scatter(df[col1], df[col2], color='blue', marker='o', label=f'{col1} (Blue Circles)')
    plt.scatter(df[col1], adjusted_y, color='red', marker='s', label=f'{col2} (Red Squares)')

    # Calculate Spearman correlation
    correlation, _ = spearmanr(df[col1].dropna(), df[col2].dropna())

    # Fit and plot a non-linear trend line for each pair
    x_new = np.linspace(min(df[col1]), max(df[col1]), 100)
    params = curve_fit(quadratic_fit, df[col1].dropna(), df[col2].dropna())[0]  # Fit parameters for the quadratic fit
    y_fit = quadratic_fit(x_new, *params)  # Apply the quadratic fit function
    plt.plot(x_new, y_fit, 'g--', label='Quadratic Fit')

    plt.title(f'Scatter Plot: {col1} vs. {col2}\nSpearman\'s rho={correlation:.2f}')
    plt.xlabel(col1)
    plt.ylabel(col2)
    plt.legend()
    plt.grid(True)

    # Save the plot to a bytes object and then to an image file
    img_data = BytesIO()
    plt.savefig(img_data, format='png')
    img_data.seek(0)
    img = Image(img_data)
    # Add image to the Excel sheet
    cell_location = f'A{1 + (index - 1) * 40}'  # Adjust cell location for each image to avoid overlap
    ws.add_image(img, cell_location)
    plt.close()

# Save the workbook with the images
wb.save(destination_file)
print(f"All scatter plots have been saved to '{destination_file}'.")


# In[38]:


pip install pandas matplotlib openpyxl scipy numpy


# In[43]:


#Spearman's Rank Correlation Coefficient instead, along with the same types of output 
#(p-value, coefficient of determination, confidence intervals, sample size, and 
#Spearman's Rank Correlation, which is more appropriate for ordinal data or when the data 
#doesn't meet the assumptions of the Pearson correlation (i.e., non-normally distributed).

import pandas as pd
from scipy.stats import spearmanr
from scipy.stats import norm
import numpy as np

# Path to the source Excel file
source_file = "/Users/vladimirvuksanovikj/Downloads/Shani Anova/EXCEL Za trudot/Power bi/korelacija od power Bi.xlsx"
# Path to the destination Excel file for results
results_file = "/Users/vladimirvuksanovikj/Downloads/Shani Anova/EXCEL Za trudot/Power bi/spearman_results.xlsx"

# Load the data
df = pd.read_excel(source_file)

# Define the column pairs for correlation calculation
pairs = [
    ("EWMA_ACWR DZ_3+4+5", "DZ 3+4+5_ACWR"),
    ("EWMA_ACWR HMLD", "HMLD_ACWR"),
    ("EWMA_ACWR TD", "Total D_ACWR"),
    ("EWMA_ACWR ACC", "ACC_ACWR"),
    ("EWMA_ACWR DEC", "DEC_ACWR")
]

results = []
for col1, col2 in pairs:
    data1, data2 = df[col1].dropna(), df[col2].dropna()
    correlation, p_value = spearmanr(data1, data2)
    r_squared = correlation**2
    sample_size = len(data1)
    
    # Confidence interval using Fisher's z transformation approximation
    fisher_z = np.arctanh(correlation)
    se = 1 / np.sqrt(sample_size - 3)
    z_critical = norm.ppf(0.975)  # 95% confidence
    ci_lower = np.tanh(fisher_z - z_critical * se)
    ci_upper = np.tanh(fisher_z + z_critical * se)

    results.append({
        'Column Pair': f'{col1} & {col2}',
        'Correlation': correlation,
        'P-value': p_value,
        '95% Confidence Interval Lower': ci_lower,
        '95% Confidence Interval Upper': ci_upper,
        'Sample Size': sample_size
    })

# Convert results to a DataFrame
result_df = pd.DataFrame(results)

# Save the results to Excel
result_df.to_excel(results_file, index=False)
print("Spearman rank correlation results have been saved to the Excel file.")


# In[45]:


#Script for Calculating Kendall's Tau Correlation Coefficient
#Kendall's Tau Calculation: This script uses kendalltau from scipy.stats, 
#which computes Kendall's Tau, a statistic that measures the ordinal association between two measured quantities. 
#It returns both the correlation coefficient and the p-value for testing the hypothesis of independence.

import pandas as pd
from scipy.stats import kendalltau

# Path to the source Excel file
source_file = "/Users/vladimirvuksanovikj/Downloads/Shani Anova/EXCEL Za trudot/Power bi/korelacija od power Bi.xlsx"
# Path to the destination Excel file for results
results_file = "/Users/vladimirvuksanovikj/Downloads/Shani Anova/EXCEL Za trudot/Power bi/kendall_tau_results.xlsx"

# Load the data
df = pd.read_excel(source_file)

# Define the column pairs for correlation calculation
pairs = [
    ("EWMA_ACWR DZ_3+4+5", "DZ 3+4+5_ACWR"),
    ("EWMA_ACWR HMLD", "HMLD_ACWR"),
    ("EWMA_ACWR TD", "Total D_ACWR"),
    ("EWMA_ACWR ACC", "ACC_ACWR"),
    ("EWMA_ACWR DEC", "DEC_ACWR")
]

results = []
for col1, col2 in pairs:
    data1, data2 = df[col1].dropna(), df[col2].dropna()
    correlation, p_value = kendalltau(data1, data2)
    results.append({
        'Column Pair': f'{col1} & {col2}',
        'Correlation': correlation,
        'P-value': p_value,
        'Sample Size': len(data1)
    })

# Convert results to a DataFrame
result_df = pd.DataFrame(results)

# Save the results to Excel
result_df.to_excel(results_file, index=False)
print("Kendall's Tau correlation results have been saved to the Excel file.")


# In[49]:


pip install dcor


# In[51]:


#Distance Correlation is a valuable statistical measure that provides a generalized and more powerful 
#alternative to the Pearson correlation coefficient. It can detect both linear and nonlinear associations 
#between variables, making it particularly useful when the relationship between the variables is not known 
#to be linear. Distance correlation takes on values between 0 and 1, where 0 indicates no correlation, and values 
#closer to 1 indicate a stronger association.

#script to calculate the Distance Correlation for specified column pairs using the Python dcor package, 
#which needs to be installed if not already available. This script will calculate the distance correlation 
#along with a significance test to determine if the observed correlation is statistically significant. 
#This significance test compares the distance correlation to what would be expected under the null hypothesis of 
#independence between the variables.

#pip install dcor //////prvo ova pred da ja run-nesh

import pandas as pd
import dcor

# Path to the source Excel file
source_file = "/Users/vladimirvuksanovikj/Downloads/Shani Anova/EXCEL Za trudot/Power bi/korelacija od power Bi.xlsx"
# Path to the destination Excel file for results
results_file = "/Users/vladimirvuksanovikj/Downloads/Shani Anova/EXCEL Za trudot/Power bi/distance_correlation_results.xlsx"

# Load the data
df = pd.read_excel(source_file)

# Define the column pairs for correlation calculation
pairs = [
    ("EWMA_ACWR DZ_3+4+5", "DZ 3+4+5_ACWR"),
    ("EWMA_ACWR HMLD", "HMLD_ACWR"),
    ("EWMA_ACWR TD", "Total D_ACWR"),
    ("EWMA_ACWR ACC", "ACC_ACWR"),
    ("EWMA_ACWR DEC", "DEC_ACWR")
]

results = []
for col1, col2 in pairs:
    data1, data2 = df[col1].dropna(), df[col2].dropna()
    distance_corr = dcor.distance_correlation(data1, data2)
    energy_distance = dcor.energy_distance(data1, data2)
    
    # Conduct the permutation test to calculate the p-value
    # It is computationally expensive, adjust `num_resamples` if needed for faster performance
    test_result = dcor.independence.distance_covariance_test(data1, data2, num_resamples=1000)
    p_value = test_result.p_value

    results.append({
        'Column Pair': f'{col1} & {col2}',
        'Distance Correlation': distance_corr,
        'P-value': p_value,
        'Energy Distance': energy_distance,
        'Sample Size': len(data1)
    })

# Convert results to a DataFrame
result_df = pd.DataFrame(results)

# Save the results to Excel
result_df.to_excel(results_file, index=False)
print("Distance correlation, p-value from permutation test, and energy distance results have been saved to the Excel file.")


# In[52]:


# SMA i EWMA normality check
#Kolmogorov-Smirnov test
#Shapiro-Wilk and Kolmogorov-Smirnov Tests:
#Skewness and Kurtosis

import pandas as pd
from scipy.stats import shapiro, kstest, skew, kurtosis
from scipy.stats import norm
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

# Load data from the Excel file
source_file = "/Users/vladimirvuksanovikj/Downloads/Shani Anova/EXCEL Za trudot/Power bi/korelacija od power Bi.xlsx"
df = pd.read_excel(source_file)

# Define the columns to test for normality
columns_to_test = [
    "EWMA_ACWR DZ_3+4+5", "DZ 3+4+5_ACWR",
    "EWMA_ACWR HMLD", "HMLD_ACWR",
    "EWMA_ACWR TD", "Total D_ACWR",
    "EWMA_ACWR ACC", "ACC_ACWR",
    "EWMA_ACWR DEC", "DEC_ACWR"
]

# Initialize a new Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Normality Test Results"

# Header for the Excel table
header = ["Column Name", "Shapiro Stat", "Shapiro P-value", "KS Stat", "KS P-value", "Skewness", "Kurtosis", "Normality"]
ws.append(header)

# Perform normality tests and measure skewness and kurtosis for each column
results = []
for column in columns_to_test:
    if column in df.columns:  # Check if the column exists in the DataFrame
        data = df[column].dropna()  # Drop NA values for the test
        shapiro_stat, shapiro_p = shapiro(data)
        ks_stat, ks_p = kstest(data, 'norm', args=(data.mean(), data.std()))
        data_skew = skew(data)
        data_kurtosis = kurtosis(data)
        is_normal = "Yes" if shapiro_p > 0.05 and ks_p > 0.05 else "No"

        # Append results
        results.append((column, shapiro_stat, shapiro_p, ks_stat, ks_p, data_skew, data_kurtosis, is_normal))

        # Write each row to the Excel sheet and apply formatting
        row = ws.max_row + 1
        ws.append([column, shapiro_stat, shapiro_p, ks_stat, ks_p, data_skew, data_kurtosis, is_normal])
        if is_normal == "No":
            for col in range(1, 9):
                ws.cell(row=row, column=col).font = Font(color="FF0000")

# Save the workbook
destination_file = "/Users/vladimirvuksanovikj/Downloads/Shani Anova/EXCEL Za trudot/Power bi/normality_check_results.xlsx"
wb.save(destination_file)
print("Extended normality test results have been saved.")


# In[53]:


#Za whiskers plot ova treba
pip install pandas matplotlib openpyxl


# In[57]:


#box-and-whisker plots

import pandas as pd
import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image

# Load data from the Excel file
source_file = "/Users/vladimirvuksanovikj/Downloads/Shani Anova/EXCEL Za trudot/Power bi/korelacija od power Bi.xlsx"
df = pd.read_excel(source_file)

# Define pairs of columns to plot side by side
column_pairs = [
    ("EWMA_ACWR DZ_3+4+5", "DZ 3+4+5_ACWR"),
    ("EWMA_ACWR HMLD", "HMLD_ACWR"),
    ("EWMA_ACWR TD", "Total D_ACWR"),
    ("EWMA_ACWR ACC", "ACC_ACWR"),
    ("EWMA_ACWR DEC", "DEC_ACWR")
]

# Initialize a new Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Side by Side Box Plots"

# Create a single figure with subplots for each pair
fig, axes = plt.subplots(nrows=len(column_pairs), ncols=2, figsize=(10, 4 * len(column_pairs)))  # 2 columns for side-by-side plotting

for i, (col1, col2) in enumerate(column_pairs):
    for j, col in enumerate([col1, col2]):
        if col in df.columns:  # Check if the column exists
            ax = axes[i, j] if len(column_pairs) > 1 else axes[j]  # Adjust for single pair case
            data = df[col].dropna()
            ax.boxplot(data, vert=True, widths=0.7, patch_artist=True)
            ax.set_title(f'Box Plot of {col}')
            ax.set_ylabel('Values')
            ax.grid(True)

plt.tight_layout()

# Save the combined plot to a BytesIO buffer
img_data = BytesIO()
plt.savefig(img_data, format='png')
img_data.seek(0)  # Go to the beginning of the BytesIO buffer

# Load this image into the Excel file
img = Image(img_data)
ws.add_image(img, 'A1')
plt.close()

# Save the workbook with the box plots
destination_file = "/Users/vladimirvuksanovikj/Downloads/Shani Anova/EXCEL Za trudot/Power bi/side_by_side_box_plots.xlsx"
wb.save(destination_file)
print("Side-by-side box-and-whisker plots have been saved to the Excel workbook.")


# In[2]:


pip install pandas numpy statsmodels scipy openpyxl matplotlib


# In[59]:


import pandas as pd
import statsmodels.api as sm
from statsmodels.stats.diagnostic import het_breuschpagan, linear_rainbow
from scipy.stats import shapiro
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Load data
df_dependent = pd.read_excel("/Users/vladimirvuksanovikj/Downloads/Shani Anova/VeljkoStoev2.xlsx",
                             usecols=["Distance 3+4+5 (m)"])
df_independent = pd.read_excel("/Users/vladimirvuksanovikj/Downloads/Shani Anova/EXCEL Za trudot/Power bi/korelacija od power Bi.xlsx",
                               usecols=["EWMA_ACWR DZ_3+4+5", "DZ 3+4+5_ACWR"])

# Ensure both DataFrames have the same length if needed
min_length = min(len(df_dependent), len(df_independent))
df_dependent = df_dependent.head(min_length)
df_independent = df_independent.head(min_length)

# Combine data into one DataFrame
df = pd.concat([df_dependent, df_independent], axis=1)

# Prepare independent variables with a constant for intercept
X = sm.add_constant(df_independent)  # adding a constant for the intercept
Y = df_dependent['Distance 3+4+5 (m)']

# Fit the regression model
model = sm.OLS(Y, X).fit()

# Getting residuals
residuals = model.resid

# Perform diagnostic tests
rainbow_stat, rainbow_p_value = linear_rainbow(model)
bp_test = het_breuschpagan(residuals, model.model.exog)
shapiro_test = shapiro(residuals)

# Create a workbook and add the results
wb = Workbook()
ws = wb.active
ws.title = "Regression Analysis"

# Write regression results and diagnostics
ws.append(["Regression Results"])
for line in model.summary().as_text().split('\n'):
    ws.append([line.strip()])

ws.append(["Diagnostic Tests"])
ws.append(["Rainbow Test Statistic", "Rainbow Test P-value"])
ws.append([rainbow_stat, rainbow_p_value])
ws.append(["Breusch-Pagan Test Statistic", "BP Test P-value", "BP Test F-statistic", "BP Test F-p-value"])
ws.append(list(bp_test))
ws.append(["Shapiro-Wilk Test Statistic", "Shapiro-Wilk P-value"])
ws.append(list(shapiro_test))

# Plotting residuals
plt.figure()
plt.scatter(model.fittedvalues, residuals)
plt.axhline(0, color='red', linestyle='--')
plt.title('Residuals vs Fitted')
plt.xlabel('Fitted values')
plt.ylabel('Residuals')
plt.show()

# Save the plot to a buffer
plt_buffer = BytesIO()
plt.savefig(plt_buffer)
plt_buffer.seek(0)

# Insert the plot into the Excel file
img = Image(plt_buffer)
ws.add_image(img, 'A50')

# Save the workbook
destination_file = "/Users/vladimirvuksanovikj/Downloads/Shani Anova/EXCEL Za trudot/Power bi/multiple_regression_analysis.xlsx"
wb.save(destination_file)
print("Multiple regression analysis results have been saved to Excel.")


# In[6]:


#Descriptivna - ama proveri za koi metriki e
import pandas as pd
from scipy.stats import shapiro, kstest, skew, kurtosis
from scipy.stats import norm

# Load your data
data = pd.read_excel('/Users/vladimirvuksanovikj/Downloads/Shani Anova/Abstrakt 2/Excel/Glavna za vo Python.xlsx')  # Make sure to update this path

# Specify the metrics to analyze
metrics = [
    'EWMA_ACWR DZ_3+4+5', 'DZ 3+4+5_ACWR', 
    'EWMA_ACWR HMLD', 'HMLD_ACWR', 'EWMA_ACWR TD', 'Total D_ACWR', 
    'EWMA_ACWR ACC', 'ACC_ACWR', 'EWMA_ACWR DEC', 'DEC_ACWR'
]


# Prepare a DataFrame to store the descriptive statistics
desc_stats = pd.DataFrame(columns=['Metric', 'Min', 'Max', 'Mean', 'SD', 'Skew', 'Kurtosis', 'Shapiro-Wilk', 'Kolmogorov-Smirnov'])

# Calculate descriptive statistics and normality tests
for metric in metrics:
    series = data[metric].dropna()
    desc_stats = desc_stats.append({
        'Metric': metric,
        'Min': series.min(),
        'Max': series.max(),
        'Mean': series.mean(),
        'SD': series.std(),
        'Skew': skew(series),
        'Kurtosis': kurtosis(series),
        'Shapiro-Wilk': shapiro(series)[1],
        'Kolmogorov-Smirnov': kstest(series, 'norm', args=(series.mean(), series.std()))[1]
    }, ignore_index=True)

# Define a function to apply the color formatting
def highlight_non_normal(val):
    color = 'orange' if val < 0.05 else None
    return f'background-color: {color}'

# Apply the color formatting
styled_desc_stats = desc_stats.style.applymap(
    highlight_non_normal, subset=['Shapiro-Wilk', 'Kolmogorov-Smirnov']
)

# Print the styled DataFrame
print(styled_desc_stats)

# To save the styled DataFrame to an Excel file, you need xlsxwriter
excel_file = '/Users/vladimirvuksanovikj/Downloads/Shani Anova/Abstrakt 2/Excel/Deskriptivna_za_SMA_EWMA.xlsx'
styled_desc_stats.to_excel(excel_file, engine='xlsxwriter')

print(f'Descriptive statistics saved to {excel_file}.')


# In[27]:


#T-test megju SMA i EWMA za site metriki, zaradi frekvencija na plus/minus rezultati
import pandas as pd
from scipy.stats import ttest_rel
import numpy as np

# Load the data from the Excel file
data = pd.read_excel('/Users/vladimirvuksanovikj/Downloads/Shani Anova/Abstrakt 2/Excel/Glavna za vo Python.xlsx')

# Define the pairs of variables
pairs = [
    ("DZ345_SMA", "DZ345_EWMA"),
    ("HMLD_SMA", "HMLD_EWMA"),
    ("TD_SMA", "TD_EWMA"),
    ("ACC_SMA", "ACC_EWMA"),
    ("DEC_SMA", "DEC_EWMA")
]

# Create an empty DataFrame to store results
results = pd.DataFrame(columns=['Variable Pair', 't-statistic', 'p-value'])

# Perform a paired t-test for each pair
for sma, ewma in pairs:
    # Ensure no missing data in pairs
    pair_data = data[[sma, ewma]].dropna()
    
    # Calculate the t-test on TWO RELATED samples
    t_stat, p_val = ttest_rel(pair_data[sma], pair_data[ewma])
    
    # Append the results
    results = results.append({
        'Variable Pair': f'{sma} / {ewma}',
        't-statistic': t_stat,
        'p-value': p_val
    }, ignore_index=True)

# Save the results to an Excel file
excel_file = '/Users/vladimirvuksanovikj/Downloads/Shani Anova/Abstrakt 2/Excel/T-Test_Frekvencii.xlsx'
with pd.ExcelWriter(excel_file) as writer:
    results.to_excel(writer, index=False, sheet_name='Paired T-Test Results')

print("Paired t-test results have been saved to", excel_file)


# In[ ]:




